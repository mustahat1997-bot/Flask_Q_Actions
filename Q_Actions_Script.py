import os
import sqlite3
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from Q_Action_Database import (
    connect_db,
    get_actions_from_db,
    get_action_type,
    get_q_action_by_site_code,
    DB_PATH
)

# 🔥 الجديد
from Points_Lookup import get_repeater_and_province_from_excel


# =========================
# NORMALIZE TEXT (IGNORE SPACES)
# =========================
def normalize_text(text):
    return re.sub(r"\s+", "", str(text)).strip().lower()


# =========================
# HELPERS
# =========================
def extract_site_code(text):
    m = re.search(r"\[(.*?)\]", str(text))
    return m.group(1) if m else ""


def extract_repeater_name(text):
    return re.sub(r"\s*\[.*?\]", "", str(text)).strip()


def parse_names(text):
    return [
        normalize_text(p)
        for p in text.replace("،", ",").replace("\n", ",").split(",")
        if p.strip()
    ]


# =========================
# STEP 1: GET ACTIONS (FROM EXCEL INSTEAD OF SELENIUM)
# =========================
def get_actions(points_text):
    results = []
    conn = connect_db()
    cursor = conn.cursor()

    points_list = [p.strip() for p in points_text.splitlines() if p.strip()]

    for point in points_list:

        # 🔍 نبحث بالـ Excel أولاً
        repeater, province = get_repeater_and_province_from_excel(point)

        if repeater == "No Repeater":
            results.append({
                "point": point,
                "repeater": "",
                "province": "",
                "q_action": "",
                "r_action": "",
                "q_action_repeater": "",
                "not_found": True
            })
            continue

        # ✅ نكمل البحث الطبيعي
        q_action, r_action = get_actions_from_db(cursor, point)

        # 🔹 🔑 نأخذ Site code فقط بين الأقواس
        site_code = extract_site_code(repeater)
        q_action_repeater = get_q_action_by_site_code(cursor, site_code)

        results.append({
            "point": point,
            "repeater": repeater,
            "province": province,
            "q_action": q_action,
            "r_action": r_action,
            "q_action_repeater": q_action_repeater,
            "not_found": False
        })

    conn.close()
    return results



# =========================
# STEP 2: APPLY RULE (UNCHANGED)
# =========================
def apply_rule_local(step1_results, rule):
    rows = []
    step1_results = [r for r in step1_results if not r.get("not_found")]

    for idx, item in enumerate(step1_results, start=1):
        point = item["point"]
        repeater_full = item["repeater"]   # 🔥 الاسم الكامل مع site code
        province = item["province"]

        site_code = extract_site_code(repeater_full)

        rows.append([
            idx,
            point,
            repeater_full,   # 🔥 نخليه كامل بدون حذف
            site_code,
            "Affiliates",
            province,
            get_action_type(rule),
            rule
        ])

    df = pd.DataFrame(rows, columns=[
        "#",
        "RepeaterName / Affiliates Name",
        "RepeaterName / Connected from",
        "Site code",
        "Traffic Area",
        "DataCenter",
        "Action Type",
        "Action needed on Repeater / Affiliates"
    ])

    base_folder = os.path.dirname(DB_PATH)
    today_str = datetime.now().strftime("%d-%m-%Y")
    filename = f"Sales order {today_str}.xlsx"
    full_path = os.path.join(base_folder, filename)

    df.to_excel(full_path, index=False)

    wb = load_workbook(full_path)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 🔥 تنسيق عام
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if row_idx == 1:
                cell.font = Font(size=12, bold=True)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.font = Font(size=12)

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # 🔥 Auto-Fit حقيقي لكل الأعمدة
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        ws.column_dimensions[column_letter].width = max_length + 3

    wb.save(full_path)
    return full_path

